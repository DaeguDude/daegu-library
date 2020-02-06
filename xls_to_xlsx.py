import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

def cvt_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    wb = Workbook()
    ws = wb.active
    
    # openpyxl starts from index 1, so rows are 20, we need to add 1 more
    # so it can run upto row 20
    # same for the ncols
    for row in range(1, nrows+1):
        for col in range(1, ncols+1):
            ws.cell(row=row, column=col, value=sheet.cell_value(row-1, col-1))

    return wb
    
xlsx_file = cvt_xls_as_xlsx('PUT-YOUR-FILE-NAME')
xlsx_file.save('NAME-TO-SAVE')
